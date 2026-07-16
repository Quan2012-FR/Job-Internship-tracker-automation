from __future__ import annotations

import sqlite3
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from src.core import database as db
from src.core.dashboard import _delete_previous_dashboards, resolve_dashboard_output_path, write_dashboard
from src.core.models import RunStats


class DashboardTests(unittest.TestCase):
    def test_resolve_dashboard_output_path_adds_date_suffix(self) -> None:
        resolved = resolve_dashboard_output_path(
            Path("engineering_job_dashboard.xlsx"),
            datetime(2026, 7, 15, 12, 0, 0),
        )
        self.assertEqual(resolved.name, "engineering_job_dashboard_20260715.xlsx")

    def test_write_dashboard_replaces_previous_family_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            temp_path = Path(tmpdir)
            base_output = temp_path / "engineering_job_dashboard.xlsx"
            previous_output = temp_path / "engineering_job_dashboard_20260714.xlsx"

            wb = Workbook()
            wb.save(previous_output)

            conn = sqlite3.connect(":memory:")
            conn.row_factory = sqlite3.Row
            db.init_db(conn)
            stats = RunStats(last_scan_date="2026-07-15T15:00:00")

            actual_output = write_dashboard(conn, base_output, stats, priority_threshold=60)

            self.assertTrue(actual_output.exists())
            self.assertFalse(previous_output.exists())
            self.assertRegex(actual_output.name, r"engineering_job_dashboard_\d{8}\.xlsx")

    def test_delete_previous_dashboards_ignores_locked_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            temp_path = Path(tmpdir)
            current_output = temp_path / "engineering_job_dashboard_20260715.xlsx"
            locked_previous_output = temp_path / "engineering_job_dashboard.xlsx"

            Workbook().save(current_output)
            Workbook().save(locked_previous_output)

            original_unlink = Path.unlink

            def locked_unlink(path: Path, missing_ok: bool = False) -> None:
                if path == locked_previous_output:
                    raise PermissionError("locked")
                original_unlink(path, missing_ok=missing_ok)

            with self.assertLogs("engineering_job_tracker.dashboard", level="WARNING") as captured:
                with patch.object(Path, "unlink", autospec=True, side_effect=locked_unlink):
                    _delete_previous_dashboards(current_output)

            self.assertTrue(locked_previous_output.exists())
            self.assertIn("Could not delete previous dashboard because it is in use", captured.output[0])


if __name__ == "__main__":
    unittest.main()
