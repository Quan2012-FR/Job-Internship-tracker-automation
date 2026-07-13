from __future__ import annotations

import sqlite3
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from src.core import database as db
from src.core.models import RunStats
from src.core.priority import APPLICATION_STATUSES


REVIEW_COLUMNS = [
    "Date Found",
    "Company",
    "Position Title",
    "Location",
    "Employment Type",
    "Job URL",
    "Application Status",
    "Notes",
]

READY_COLUMNS = [
    "Company",
    "Position Title",
    "Location",
    "Employment Type",
    "Date Found",
    "Application Deadline",
    "Days Remaining",
    "Priority Score",
    "Job URL",
    "Application Status",
    "Notes",
]

THIS_WEEK_COLUMNS = [
    "Company",
    "Position Title",
    "Location",
    "Days Remaining",
    "Priority Score",
    "Job URL",
    "Application Status",
]

ACTIVE_COLUMNS = [
    "Company",
    "Position Title",
    "Location",
    "Employment Type",
    "Date First Seen",
    "Last Seen",
    "Application Deadline",
    "Days Remaining",
    "Priority Score",
    "Job URL",
    "Application Status",
]

RED_FILL = PatternFill("solid", fgColor="F4CCCC")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
GREEN_FILL = PatternFill("solid", fgColor="D9EAD3")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")


def _fill_review_sheet(sheet, rows) -> None:
    sheet.append(REVIEW_COLUMNS)
    for row in rows:
        sheet.append(
            [
                _date_only(row["date_found"]),
                row["company"],
                row["title"],
                row["location"],
                row["employment_type"],
                row["url"],
                row["application_status"] or "Not Started",
                row["notes"] or "",
            ]
        )


def _fill_ready_sheet(sheet, rows) -> None:
    sheet.append(READY_COLUMNS)
    for row in rows:
        sheet.append(
            [
                row["company"],
                row["title"],
                row["location"],
                row["employment_type"],
                _date_only(row["date_found"]),
                row["application_deadline"] or "Unknown",
                _days_remaining(row),
                row["priority_score"],
                row["url"],
                row["application_status"] or "Not Started",
                row["notes"] or "",
            ]
        )


def _fill_this_week_sheet(sheet, rows) -> None:
    sheet.append(THIS_WEEK_COLUMNS)
    for row in rows:
        sheet.append(
            [
                row["company"],
                row["title"],
                row["location"],
                _days_remaining(row),
                row["priority_score"],
                row["url"],
                row["application_status"] or "Not Started",
            ]
        )


def _fill_active_sheet(sheet, rows) -> None:
    sheet.append(ACTIVE_COLUMNS)
    for row in rows:
        sheet.append(
            [
                row["company"],
                row["title"],
                row["location"],
                row["employment_type"],
                _date_only(row["first_seen"]),
                _date_only(row["last_seen"]),
                row["application_deadline"] or "Unknown",
                _days_remaining(row),
                row["priority_score"],
                row["url"],
                row["application_status"] or "Not Started",
            ]
        )


def write_dashboard(conn: sqlite3.Connection, output_path: Path, stats: RunStats, priority_threshold: int) -> None:
    _sync_application_tracking(conn, output_path)

    wb = Workbook()

    ws_this_week = wb.active
    ws_this_week.title = "THIS WEEK"
    ws_ready = wb.create_sheet("READY TO APPLY")
    ws_review = wb.create_sheet("NEEDS REVIEW")
    ws_new = wb.create_sheet("NEW JOBS")
    ws_active = wb.create_sheet("ALL ACTIVE JOBS")
    ws_stats = wb.create_sheet("STATISTICS")

    last_7_days = datetime.now() - timedelta(days=7)
    this_week_rows = db.get_this_week_jobs(conn, last_7_days)
    ready_rows = db.get_ready_to_apply_jobs(conn, priority_threshold)
    review_rows = db.get_jobs_found_since(conn, last_7_days)
    new_rows = db.get_new_jobs_all_time(conn)
    active_rows = db.get_active_jobs(conn)

    _fill_this_week_sheet(ws_this_week, this_week_rows)
    _fill_ready_sheet(ws_ready, ready_rows)
    _fill_review_sheet(ws_review, review_rows)
    _fill_review_sheet(ws_new, new_rows)
    _fill_active_sheet(ws_active, active_rows)

    ws_stats.append(["Metric", "Value"])
    ws_stats.append(["Companies Checked", stats.companies_checked])
    ws_stats.append(["Jobs Found This Run", stats.jobs_found_this_run])
    ws_stats.append(["New Jobs This Run", stats.new_jobs_this_run])
    ws_stats.append(["Total Active Jobs", stats.total_active_jobs])
    ws_stats.append(["Last Scan Date", stats.last_scan_date])
    ws_stats.append(["Priority Threshold", priority_threshold])

    for sheet in wb.worksheets:
        _style_sheet(sheet)
        _apply_hyperlinks(sheet)
        _apply_status_validation(sheet)
        _apply_deadline_colors(sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _sync_application_tracking(conn: sqlite3.Connection, output_path: Path) -> None:
    if not output_path.exists():
        return

    wb = load_workbook(output_path)
    try:
        tracking_by_url: dict[str, tuple[str, str, bool]] = {}
        for sheet in wb.worksheets:
            headers = _header_map(sheet)
            url_col = headers.get("Job URL")
            status_col = headers.get("Application Status") or headers.get("Status")
            notes_col = headers.get("Notes")
            if not url_col or not status_col:
                continue
            for row_idx in range(2, sheet.max_row + 1):
                url = _cell_text(sheet.cell(row=row_idx, column=url_col).value)
                status = _cell_text(sheet.cell(row=row_idx, column=status_col).value)
                if not url or not status:
                    continue
                if status == "Not Reviewed":
                    status = "Not Started"
                notes = _cell_text(sheet.cell(row=row_idx, column=notes_col).value) if notes_col else ""
                user_value = status != "Not Started" or bool(notes)
                existing = tracking_by_url.get(url)
                if existing is None or user_value or not existing[2]:
                    tracking_by_url[url] = (status, notes, user_value)
        for url, (status, notes, _) in tracking_by_url.items():
            db.update_application_tracking(conn, url, status, notes)
        conn.commit()
    finally:
        wb.close()


def _style_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_length + 2, 12), 60)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def _apply_hyperlinks(sheet) -> None:
    url_col = _header_map(sheet).get("Job URL")
    if not url_col:
        return
    for row_idx in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row_idx, column=url_col)
        if cell.value:
            cell.hyperlink = str(cell.value)
            cell.style = "Hyperlink"


def _apply_status_validation(sheet) -> None:
    headers = _header_map(sheet)
    status_col = headers.get("Application Status") or headers.get("Status")
    if not status_col:
        return
    status_formula = ",".join(APPLICATION_STATUSES)
    validation = DataValidation(type="list", formula1=f'"{status_formula}"', allow_blank=False)
    sheet.add_data_validation(validation)
    validation.add(f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}10000")


def _apply_deadline_colors(sheet) -> None:
    headers = _header_map(sheet)
    days_col = headers.get("Days Remaining")
    if not days_col:
        return
    priority_col = headers.get("Priority Score")
    for row_idx in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row_idx, column=days_col).value
        if not isinstance(value, int):
            continue
        fill = _deadline_fill(value)
        sheet.cell(row=row_idx, column=days_col).fill = fill
        if priority_col:
            sheet.cell(row=row_idx, column=priority_col).fill = fill


def _deadline_fill(days_remaining: int) -> PatternFill:
    if days_remaining <= 7:
        return RED_FILL
    if days_remaining <= 14:
        return YELLOW_FILL
    return GREEN_FILL


def _header_map(sheet) -> dict[str, int]:
    return {str(cell.value): cell.column for cell in sheet[1] if cell.value}


def _date_only(value: str) -> str:
    text = _cell_text(value)
    if "T" in text:
        return text.split("T", 1)[0]
    return text


def _days_remaining(row) -> int | str:
    value = row["days_remaining"]
    return value if value is not None else "Unknown"


def _cell_text(value) -> str:
    return str(value or "").strip()
