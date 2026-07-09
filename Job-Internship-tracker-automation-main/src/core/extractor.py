from __future__ import annotations

from collections.abc import Iterable
from typing import Optional

from openpyxl import load_workbook

from config import WorkbookMapping
from src.core.models import CompanyTarget
from src.core.utils import extract_first_url, fallback_company_from_url, normalize_text


COMPANY_HEADER_CANDIDATES = {
    "company",
    "company name",
    "employer",
    "organization",
}

URL_HEADER_CANDIDATES = {
    "career url",
    "careers url",
    "careers page",
    "careers",
    "job board",
    "job page",
    "apply where",
    "application url",
    "job url",
    "url",
}


def _normalize_header(value: object) -> str:
    return normalize_text(str(value) if value is not None else "").lower()


def _find_header_indexes(headers: Iterable[object], mapping: WorkbookMapping) -> tuple[Optional[int], Optional[int]]:
    company_idx: Optional[int] = None
    url_idx: Optional[int] = None

    normalized = [_normalize_header(v) for v in headers]
    if mapping.company_column:
        expected = mapping.company_column.strip().lower()
        if expected in normalized:
            company_idx = normalized.index(expected)
    if mapping.careers_url_column:
        expected = mapping.careers_url_column.strip().lower()
        if expected in normalized:
            url_idx = normalized.index(expected)

    if company_idx is None:
        for i, h in enumerate(normalized):
            if h in COMPANY_HEADER_CANDIDATES:
                company_idx = i
                break

    if url_idx is None:
        for i, h in enumerate(normalized):
            if h in URL_HEADER_CANDIDATES:
                url_idx = i
                break

    if company_idx is None and normalized:
        company_idx = 0
    return company_idx, url_idx


def extract_company_targets(path: str, mapping: WorkbookMapping) -> list[CompanyTarget]:
    wb = load_workbook(path, read_only=True, data_only=True)
    targets: list[CompanyTarget] = []
    seen_pairs: set[tuple[str, str]] = set()

    for ws in wb.worksheets:
        if mapping.sheet_name and ws.title != mapping.sheet_name:
            continue

        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            continue

        company_idx, url_idx = _find_header_indexes(header_row, mapping)
        if company_idx is None or url_idx is None:
            continue

        header_names = [normalize_text(str(v) if v is not None else "") for v in header_row]
        for row_number, row in enumerate(rows, start=2):
            company_raw = row[company_idx] if company_idx < len(row) else None
            url_raw = row[url_idx] if url_idx < len(row) else None

            url = extract_first_url(str(url_raw) if url_raw is not None else "")
            if not url:
                continue

            company = normalize_text(str(company_raw) if company_raw is not None else "")
            if not company:
                company = fallback_company_from_url(url)

            key = (company.lower(), url.lower())
            if key in seen_pairs:
                continue
            seen_pairs.add(key)

            metadata: dict[str, str] = {}
            for i, value in enumerate(row):
                if i >= len(header_names):
                    continue
                header = header_names[i]
                if not header or i in (company_idx, url_idx):
                    continue
                text = normalize_text(str(value) if value is not None else "")
                if text:
                    metadata[header] = text

            targets.append(
                CompanyTarget(
                    company=company,
                    careers_url=url,
                    source_sheet=ws.title,
                    source_row=row_number,
                    metadata=metadata,
                )
            )
    return targets
